from typing import BinaryIO
from decimal import Decimal, InvalidOperation

import openpyxl
from sqlalchemy.orm import Session

from app.models.student_indicator import StudentIndicator
from app.models.audit_log import AuditLog
from app.schemas.import_schema import (
    ImportError,
    ImportChange,
    ImportPendingRow,
    ImportPreviewResponse,
)

EXPECTED_HEADERS = [
    "PERIODO",
    "ESTADO",
    "VINCULACION",
    "BRA",
    "SEMESTRES_CURSADOS",
    "PENSUM_APROBADO_PCT",
    "PROMEDIO_ACUMULADO",
    "COLEGIO_ORIGEN",
    "SEXO",
    "ESTRATO",
    "TESIS_ESTADO",
    "TESIS_NOTA",
    "ACOMPANAMIENTO_BRA",
    "PRACTICA_PROFESIONAL",
]

FIELD_MAP = {
    "PERIODO": "periodo",
    "ESTADO": "estado",
    "VINCULACION": "vinculacion",
    "BRA": "bra",
    "SEMESTRES_CURSADOS": "semestres_cursados",
    "PENSUM_APROBADO_PCT": "pensum_aprobado_pct",
    "PROMEDIO_ACUMULADO": "promedio_acumulado",
    "COLEGIO_ORIGEN": "colegio_origen",
    "SEXO": "sexo",
    "ESTRATO": "estrato",
    "TESIS_ESTADO": "tesis_estado",
    "TESIS_NOTA": "tesis_nota",
    "ACOMPANAMIENTO_BRA": "acompanamiento_bra",
    "PRACTICA_PROFESIONAL": "practica_profesional",
}


def parse_excel(file: BinaryIO) -> tuple[list[dict], list[ImportError]]:
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    errors: list[ImportError] = []
    rows: list[dict] = []

    headers = [cell.value for cell in ws[1]]
    print(f"DEBUG: Headers found in Excel: {headers}")

    if headers != EXPECTED_HEADERS:
        missing = [h for h in EXPECTED_HEADERS if h not in headers]
        if missing:
            errors.append(
                ImportError(row=1, field="headers", message=f"Faltan columnas: {', '.join(missing)}")
            )
            wb.close()
            return [], errors

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        record = {"_row": row_idx}
        for col_idx, header in enumerate(headers):
            value = row[col_idx]
            record[FIELD_MAP[header]] = value
        rows.append(record)

    wb.close()
    return rows, errors


def _validate_row(row: dict) -> list[ImportError]:
    errors: list[ImportError] = []
    row_num = row["_row"]

    if not row.get("periodo"):
        errors.append(ImportError(row=row_num, field="periodo", message="Período requerido"))

    if not row.get("estado"):
        errors.append(ImportError(row=row_num, field="estado", message="Estado requerido"))

    if not row.get("vinculacion"):
        errors.append(ImportError(row=row_num, field="vinculacion", message="Vinculación requerida"))

    bra_val = row.get("bra")
    if bra_val is not None:
        try:
            int(bra_val)
        except (ValueError, TypeError):
            errors.append(ImportError(row=row_num, field="bra", message="BRA debe ser un entero"))

    if not row.get("semestres_cursados"):
        errors.append(
            ImportError(row=row_num, field="semestres_cursados", message="Semestres cursados requerido")
        )

    try:
        val = row.get("pensum_aprobado_pct")
        if val is not None:
            Decimal(str(val))
    except (InvalidOperation, ValueError):
        errors.append(
            ImportError(row=row_num, field="pensum_aprobado_pct", message="Porcentaje inválido")
        )

    try:
        val = row.get("promedio_acumulado")
        if val is not None:
            d = Decimal(str(val))
            if d < 0 or d > 5:
                errors.append(
                    ImportError(row=row_num, field="promedio_acumulado", message="Promedio debe ser 0-5")
                )
    except (InvalidOperation, ValueError):
        errors.append(
            ImportError(row=row_num, field="promedio_acumulado", message="Promedio inválido")
        )

    sexo = row.get("sexo")
    if sexo and sexo not in ("M", "F"):
        errors.append(ImportError(row=row_num, field="sexo", message="Sexo debe ser M o F"))

    estrato = row.get("estrato")
    if estrato is not None:
        try:
            e = int(estrato)
            if e < 1 or e > 6:
                errors.append(
                    ImportError(row=row_num, field="estrato", message="Estrato debe ser 1-6")
                )
        except (ValueError, TypeError):
            errors.append(ImportError(row=row_num, field="estrato", message="Estrato inválido"))

    if not row.get("tesis_estado"):
        errors.append(ImportError(row=row_num, field="tesis_estado", message="Estado de tesis requerido"))

    return errors


def _normalize_row(row: dict) -> dict:
    bra_val = row.get("bra")
    try:
        bra = int(bra_val) if bra_val is not None else 0
    except (ValueError, TypeError):
        bra = 0

    acomp = str(row.get("acompanamiento_bra", "NO")).upper()
    acompanamiento_bra = acomp in ("1", "TRUE", "YES", "SI", "SÍ")

    prac = str(row.get("practica_profesional", "NO")).upper()
    practica_profesional = prac in ("1", "TRUE", "YES", "SI", "SÍ")

    tesis_nota = row.get("tesis_nota")
    if tesis_nota is not None:
        try:
            tesis_nota = Decimal(str(tesis_nota))
        except (InvalidOperation, ValueError):
            tesis_nota = None

    pensum = row.get("pensum_aprobado_pct")
    if pensum is not None:
        try:
            pensum = Decimal(str(pensum))
        except (InvalidOperation, ValueError):
            pensum = Decimal("0")

    promedio = row.get("promedio_acumulado")
    if promedio is not None:
        try:
            promedio = Decimal(str(promedio))
        except (InvalidOperation, ValueError):
            promedio = Decimal("0")

    estrato = row.get("estrato")
    if estrato is not None:
        try:
            estrato = int(estrato)
        except (ValueError, TypeError):
            estrato = 0

    return {
        "periodo": str(row.get("periodo", "")),
        "estado": str(row.get("estado", "")),
        "vinculacion": str(row.get("vinculacion", "")),
        "bra": bra,
        "semestres_cursados": str(row.get("semestres_cursados", "")),
        "pensum_aprobado_pct": pensum,
        "promedio_acumulado": promedio,
        "colegio_origen": str(row.get("colegio_origen", "")),
        "sexo": str(row.get("sexo", "")),
        "estrato": estrato,
        "tesis_estado": str(row.get("tesis_estado", "")),
        "tesis_nota": tesis_nota,
        "acompanamiento_bra": acompanamiento_bra,
        "practica_profesional": practica_profesional,
    }


def _detect_changes(old: dict, new: dict) -> list[str]:
    changed = []
    for key in new:
        if key == "_row":
            continue
        old_val = old.get(key)
        new_val = new[key]
        if str(old_val) != str(new_val):
            changed.append(key)
    return changed


def preview_import(db: Session, file: BinaryIO, filename: str) -> ImportPreviewResponse:
    print(f"DEBUG: Starting import preview for file: {filename}")
    rows, parse_errors = parse_excel(file)

    all_errors = list(parse_errors)
    to_create = 0
    to_update = 0
    unchanged = 0
    sample_changes: list[ImportChange] = []
    pending_data: list[ImportPendingRow] = []

    for row in rows:
        validation_errors = _validate_row(row)
        if validation_errors:
            all_errors.extend(validation_errors)
            continue

        normalized = _normalize_row(row)
        row_num = row["_row"]

        existing = (
            db.query(StudentIndicator)
            .filter(StudentIndicator.periodo == normalized["periodo"])
            .all()
        )

        if not existing:
            to_create += 1
            pending_data.append(ImportPendingRow(action="create", row=row_num, data=normalized))
            if len(sample_changes) < 10:
                sample_changes.append(
                    ImportChange(
                        row=row_num,
                        periodo=normalized["periodo"],
                        action="create",
                        fields_changed=list(normalized.keys()),
                        new_values=normalized,
                    )
                )
        else:
            match = existing[0]
            old_dict = {
                "periodo": match.periodo,
                "estado": match.estado,
                "vinculacion": match.vinculacion,
                "bra": match.bra,
                "semestres_cursados": match.semestres_cursados,
                "pensum_aprobado_pct": float(match.pensum_aprobado_pct) if match.pensum_aprobado_pct else 0,
                "promedio_acumulado": float(match.promedio_acumulado) if match.promedio_acumulado else 0,
                "colegio_origen": match.colegio_origen,
                "sexo": match.sexo,
                "estrato": match.estrato,
                "tesis_estado": match.tesis_estado,
                "tesis_nota": float(match.tesis_nota) if match.tesis_nota else None,
                "acompanamiento_bra": match.acompanamiento_bra,
                "practica_profesional": match.practica_profesional,
            }

            fields_changed = _detect_changes(old_dict, normalized)
            if fields_changed:
                to_update += 1
                pending_data.append(
                    ImportPendingRow(
                        action="update",
                        row=row_num,
                        data=normalized,
                        record_id=match.id,
                        old=old_dict,
                    )
                )
                if len(sample_changes) < 10:
                    sample_changes.append(
                        ImportChange(
                            row=row_num,
                            periodo=normalized["periodo"],
                            action="update",
                            fields_changed=fields_changed,
                            old_values={k: old_dict[k] for k in fields_changed},
                            new_values={k: normalized[k] for k in fields_changed},
                        )
                    )
            else:
                unchanged += 1

    return ImportPreviewResponse(
        filename=filename,
        total_rows=len(rows),
        to_create=to_create,
        to_update=to_update,
        unchanged=unchanged,
        errors=all_errors,
        sample_changes=sample_changes,
        pending_rows=pending_data,
    )


def execute_import(db: Session, rows: list[ImportPendingRow], user_id: int, username: str) -> dict:
    created = 0
    updated = 0

    for item in rows:
        action = item.action
        data = item.data

        if action == "create":
            new_record = StudentIndicator(**data)
            db.add(new_record)
            db.flush()

            log = AuditLog(
                user_id=user_id,
                username=username,
                action="IMPORT_CREATE",
                entity="student_indicators",
                record_id=new_record.id,
                old_values=None,
                new_values=data,
                import_batch="direct",
            )
            db.add(log)
            created += 1

        elif action == "update":
            record_id = item.record_id
            old_values = item.old
            record = db.query(StudentIndicator).filter(StudentIndicator.id == record_id).first()
            if record:
                for key, value in data.items():
                    setattr(record, key, value)

                log = AuditLog(
                    user_id=user_id,
                    username=username,
                    action="IMPORT_UPDATE",
                    entity="student_indicators",
                    record_id=record_id,
                    old_values=old_values,
                    new_values=data,
                    import_batch="direct",
                )
                db.add(log)
                updated += 1

    db.commit()

    return {
        "success": True,
        "created": created,
        "updated": updated,
        "message": f"Importación completada: {created} creados, {updated} actualizados.",
    }
